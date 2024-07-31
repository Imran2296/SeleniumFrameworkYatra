import time

from select import select
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == colName:
          Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:/Users/191485/Downloads/download (1).xlsx"
fruit_name = "Apple"
newValue = "990"
service_obj = Service("C:\\Webdrivers\\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj)
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()

#edit the excel with updated value
update_excel_data(file_path, fruit_name, "price",newValue)

#since uploading is done through local windows, just clicking on upload button cannot upload it
#so the html should have type='file' attribute
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path) #pass the path of the file here

wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")#this is locator of success message which
#vanishes after success message dissappears.
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

#dynamically updating prices of fruits with xpth
price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
assert actual_price == newValue